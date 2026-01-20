import csv
import hashlib
import os
import textwrap
from zipfile import ZipFile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

XLSX_PATH = "AVC_MASTER_AXIS.xlsx"
ZIP_PATH = "AVC_MASTER_AXIS_PACK.zip"
README_PATH = "README_MASTER_AXIS.txt"

CSV_TEMPLATES = {
    "master_labor_axis.csv": [
        [
            "Activity/Asset",
            "Track",
            "Status",
            "Date Range.start",
            "Date Range.end",
            "Market Role",
            "Proven Hours",
            "Fixation Flag",
            "Evidence ID",
            "Notes",
        ],
        [
            "Drafting 501c3 Bylaws",
            "Track B (Architectural)",
            "Unpaid / Extracted",
            "2020-04-01",
            "2020-05-01",
            "Non-Profit Consultant",
            "46",
            "TRUE",
            "EV-2021-019",
            "Director requested virtual pivot; bylaws drafted and shared.",
        ],
    ],
    "rates.csv": [
        [
            "Role",
            "base_rate_low",
            "base_rate_high",
            "base_rate_max",
            "risk_multiplier",
            "valuation_notes",
        ],
        ["Non-Profit Consultant", "150", "250", "300", "1.1", "Emergency pivot rate; locked gate"],
    ],
    "evidence_vault.csv": [
        ["evidence_id", "file_name", "date", "source", "description", "hash_optional"],
        [
            "EV-2021-019",
            "email_director_virtualsetup.eml",
            "2020-03-28",
            "Email",
            "Director requested virtual setup.",
            "",
        ],
    ],
    "asset_log.csv": [
        [
            "asset_id",
            "asset_type",
            "created_date",
            "creator",
            "first_use",
            "reuse_count",
            "Market License Fee",
            "monetized_flag",
            "evidence_ref",
        ],
        ["VID-XYZ-2021", "Video", "2021-05-05", "Allison", "2021-05-06", "12", "100", "TRUE", "EV-2021-019"],
    ],
}


def sha256_of_file(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(8192), b""):
            digest.update(chunk)
    return digest.hexdigest()


def add_sheet(workbook: Workbook, title: str, headers: list[str], fill_color: str) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor=fill_color)
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index)
        cell.font = header_font
        cell.fill = header_fill
        sheet.column_dimensions[chr(64 + col_index)].width = max(18, len(header) + 2)


def build_workbook() -> Workbook:
    workbook = Workbook()
    add_sheet(
        workbook,
        "CHRONOLOGY_AXIS",
        [
            "record_id",
            "start_date",
            "end_date",
            "year",
            "month",
            "role_at_time",
            "activity",
            "location",
            "paid_flag",
            "evidence_type",
            "notes",
        ],
        "DDDDDD",
    )
    add_sheet(workbook, "BOUNDARIES_CONSENT", ["boundary_id", "year", "scope", "statement", "evidence_ref"], "D9E8FB")
    add_sheet(workbook, "ROLE_ADDITIONS", ["role_id", "role_name", "first_date", "trigger_event", "evidence_ref"], "DFF0D8")
    add_sheet(
        workbook,
        "LABOR_HOURS_PROVEN",
        ["labor_id", "date", "role", "activity", "hours_proven", "proof_type", "proof_ref"],
        "FFF2CC",
    )
    add_sheet(
        workbook,
        "VALUATION_SLOTS_LOCKED",
        [
            "role",
            "base_rate_low",
            "base_rate_high",
            "base_rate_max",
            "risk_multiplier",
            "effective_rate",
            "valuation_notes",
        ],
        "FFE5CC",
    )
    add_sheet(
        workbook,
        "ASSET_FIXATION_LOG",
        [
            "asset_id",
            "asset_type",
            "created_date",
            "creator",
            "first_use",
            "reuse_count",
            "Market License Fee",
            "monetized_flag",
            "evidence_ref",
        ],
        "E6D9FF",
    )
    add_sheet(
        workbook,
        "EVIDENCE_INDEX",
        ["evidence_id", "file_name", "date", "source", "description", "hash_optional", "attached_path"],
        "333333",
    )
    add_sheet(
        workbook,
        "RECONCILIATION_AUTO",
        ["period", "hours_total", "paid_amount", "unpaid_hours", "effective_rate"],
        "D9E8FB",
    )

    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    workbook["VALUATION_SLOTS_LOCKED"].append(
        ["Non-Profit Consultant", 150, 250, 300, 1.1, 300, "Emergency pivot rate"]
    )
    return workbook


def write_csv_templates() -> None:
    for filename, rows in CSV_TEMPLATES.items():
        with open(filename, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerows(rows)


def write_evidence_csv(evidence_dir: str) -> str | None:
    if not os.path.isdir(evidence_dir):
        return None

    rows: list[list[str]] = []
    for root, _, files in os.walk(evidence_dir):
        for filename in files:
            path = os.path.join(root, filename)
            digest = sha256_of_file(path)
            evidence_id = f"EV-{int(os.path.getmtime(path))}-{abs(hash(filename)) % 100000}"
            rows.append([evidence_id, filename, "", "Local", "Auto-imported evidence", digest, path])

    if not rows:
        return None

    evidence_csv = "evidence_vault.csv"
    with open(evidence_csv, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["evidence_id", "file_name", "date", "source", "description", "hash_optional", "attached_path"])
        writer.writerows(rows)
    return evidence_csv


def write_readme() -> None:
    content = textwrap.dedent(
        """
        AVC MASTER AXIS PACK - SYSTEM OPERATING PROCEDURES

        ORDER OF ENTRY:
        1) CHRONOLOGY_AXIS
        2) ROLE_ADDITIONS
        3) LABOR_HOURS_PROVEN
        4) ASSET_FIXATION_LOG
        5) EVIDENCE_INDEX (attach files; compute SHA-256 for hash_optional)

        THE GOLDEN RULES:
        - NO ESTIMATES: leave unknown durations blank.
        - NO BACKFILLING.
        - EVIDENCE MANDATORY: map to EVIDENCE_INDEX.
        - VALUATION GATE: do not compute effective rates until Labor Hours sheet is frozen.

        Forensic block template:
        HASH: [SHA-256]
        AUTHOR: Allison
        ORIGIN: [Original Filename]
        MODIFIED: [Last Date of Extraction]
        """
    ).strip()
    with open(README_PATH, "w", encoding="utf-8") as handle:
        handle.write(content)


def build_pack(evidence_dir: str = "evidence") -> None:
    workbook = build_workbook()
    workbook.save(XLSX_PATH)
    write_csv_templates()
    evidence_csv = write_evidence_csv(evidence_dir)
    write_readme()

    with ZipFile(ZIP_PATH, "w") as zip_handle:
        zip_handle.write(XLSX_PATH)
        zip_handle.write(README_PATH)
        for filename in CSV_TEMPLATES.keys():
            zip_handle.write(filename)
        if evidence_csv:
            zip_handle.write(evidence_csv)


if __name__ == "__main__":
    build_pack()
