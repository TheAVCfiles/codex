"""Generate the AVC economic damages document pack.

This script:
1) Populates the Timeline_Spine sheet in the provided Excel workbook.
2) Generates three PDFs (damages exhibit, causation map, stress test).
3) Zips all generated artifacts into a single archive.

Dependencies:
- openpyxl
- reportlab
"""
from __future__ import annotations

import argparse
import zipfile
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_BASE_NAME = "AVC_Economic_Damages_Pack_NYC_v1"


def append_timeline_row(spine_path: Path, output_path: Path) -> None:
    """Append the requested timeline row and save to ``output_path``."""

    workbook = load_workbook(spine_path)
    if "Timeline_Spine" not in workbook.sheetnames:
        raise KeyError("Timeline_Spine sheet not found in workbook.")

    timeline = workbook["Timeline_Spine"]
    timeline.append(
        [
            "2020–2023",
            "Interrupted professional education pathways (Yoga + NYU Yellowbrick)",
            "AVC; NYU Yellowbrick; NYC studios",
            "Credential pursuit / pipeline access",
            "Emails, scholarship records, absence of certification",
            "NYC Metro",
            "Delay-based economic suppression",
        ]
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def create_pdf(path: Path, title: str, body_lines: Iterable[str]) -> None:
    """Create a simple PDF with a title and body lines."""

    styles = getSampleStyleSheet()
    story = [Paragraph(f"<b>{title}</b>", styles["Title"])]
    for line in body_lines:
        story.append(Paragraph(line, styles["BodyText"]))

    doc = SimpleDocTemplate(str(path), pagesize=LETTER)
    doc.build(story)


def build_documents(output_dir: Path, base_name: str) -> dict[str, Path]:
    """Build the three PDF documents and return their paths."""

    damages_pdf = output_dir / f"{base_name}_Damages_Exhibit.pdf"
    causation_pdf = output_dir / f"{base_name}_Causation_Map.pdf"
    stress_pdf = output_dir / f"{base_name}_Stress_Test.pdf"

    create_pdf(
        damages_pdf,
        "Economic Damages — NYC Metro (Mid-Tier)",
        [
            "Yoga Certification Delay: $214,500 (3 years @ $71,500/year)",
            "NYU Yellowbrick Stage & Screen Delay: $315,000 (3 years @ $105,000/year)",
            "Subtotal: $529,500",
            "NYC Opportunity Cost Multiplier (20%): $105,900",
            "<b>Total Estimated Economic Damages: $635,400</b>",
        ],
    )

    create_pdf(
        causation_pdf,
        "Causation Map — Event to Economic Harm",
        [
            "Event: Interruption of active educational pathways.",
            "Effect: Credentials not obtained on expected timeline.",
            "Result: Exclusion from credential-gated NYC markets.",
            "Damage: Quantifiable loss of earning capacity.",
        ],
    )

    create_pdf(
        stress_pdf,
        "Stress Test & Defensibility Analysis",
        [
            "Rates are mid-tier NYC market averages.",
            "Damages reflect delay, not speculative success.",
            "20% opportunity multiplier is within accepted NYC ranges.",
            "Parallel income tracks are standard in NYC creative economies.",
        ],
    )

    return {
        "damages": damages_pdf,
        "causation": causation_pdf,
        "stress": stress_pdf,
    }


def create_zip(paths: Iterable[Path], zip_path: Path) -> Path:
    """Zip the provided files and return the archive path."""

    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for file_path in paths:
            archive.write(file_path, file_path.name)
    return zip_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Create the AVC economic damages document pack.")
    parser.add_argument(
        "--spine-path",
        type=Path,
        required=True,
        help="Path to the existing AVC_Archival_Intelligence_Spine workbook.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=REPO_ROOT / "dist" / DEFAULT_BASE_NAME,
        help="Directory where the generated files should be written.",
    )
    parser.add_argument(
        "--base-name",
        type=str,
        default=DEFAULT_BASE_NAME,
        help="Base name used for generated files and the zip archive.",
    )
    args = parser.parse_args()

    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    populated_spine_path = output_dir / f"{args.base_name}_Spine.xlsx"
    append_timeline_row(args.spine_path, populated_spine_path)

    pdf_paths = build_documents(output_dir, args.base_name)

    zip_path = output_dir / f"{args.base_name}.zip"
    create_zip([populated_spine_path, *pdf_paths.values()], zip_path)

    print(f"Populated spine workbook: {populated_spine_path}")
    print(f"Damages exhibit: {pdf_paths['damages']}")
    print(f"Causation map: {pdf_paths['causation']}")
    print(f"Stress test: {pdf_paths['stress']}")
    print(f"Zip archive: {zip_path}")


if __name__ == "__main__":
    main()
